from VCard import Bot
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Dados.xlsx') 
ws = wb['Sheet1']
numColaboradores = 163
for linha in range(2,numColaboradores+2):
    firstName = ws["A"+str(linha)]._value
    lastName = ws["B"+str(linha)]._value
    fullName = f"{firstName} {lastName}"
    title = ws["G"+str(linha)]._value
    cel = f'+55{ws["H"+str(linha)]._value}'
    email = ws["I"+str(linha)]._value
    address = f'{ws["K"+str(linha)]._value}, {(ws["L"+str(linha)]._value).replace("ã","a")}, {ws["M"+str(linha)]._value}- {ws["J"+str(linha)]._value}'
    Bot(dados={
        "Name":fullName,
        "FirstName":firstName,
        "LastName":lastName,
        "Title":title,
        "Address":address,
        "Celular":cel,
        "Email":email,
    }).create()
